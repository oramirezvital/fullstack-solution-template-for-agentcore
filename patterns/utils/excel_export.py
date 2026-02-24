"""
Excel export utilities for portfolio data.

This module provides functions to generate Excel files with portfolio
information including transactions, performance metrics, and forecasts.
"""

from datetime import datetime
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import base64
from io import BytesIO


def create_portfolio_excel(
    portfolio_data: Dict[str, Any],
    user_email: str = "user"
) -> str:
    """
    Generate an Excel file with portfolio details.
    
    Creates a multi-sheet Excel workbook containing:
    - Summary: Overall portfolio performance
    - Transactions: Detailed transaction history
    - Positions: Current active positions
    - Forecasts: Forecast vs. actual comparison
    
    Args:
        portfolio_data: Portfolio data from InvestmentTracker
        user_email: User email for file metadata
        
    Returns:
        Base64-encoded Excel file content
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    _create_summary_sheet(wb, portfolio_data)
    _create_transactions_sheet(wb, portfolio_data)
    _create_positions_sheet(wb, portfolio_data)
    _create_forecasts_sheet(wb, portfolio_data)
    
    # Set workbook properties
    wb.properties.creator = "Stock Market Data Agent"
    wb.properties.title = f"Portfolio Export - {user_email}"
    wb.properties.created = datetime.now()
    
    # Save to BytesIO and encode as base64
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return base64.b64encode(buffer.read()).decode('utf-8')


def _create_summary_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create summary sheet with overall portfolio metrics."""
    ws = wb.create_sheet("Summary", 0)
    
    # Header styling
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws['A1'] = "Portfolio Performance Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Export date
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    # Metrics
    row = 4
    metrics = [
        ("Total Positions", data.get('total_positions', 0)),
        ("Total Invested", f"${data.get('total_invested', 0):,.2f}"),
        ("Current Value", f"${data.get('current_value', 0):,.2f}"),
        ("Total Gain/Loss", f"${data.get('total_gain_loss', 0):,.2f}"),
        ("Gain/Loss %", f"{data.get('total_gain_loss_pct', 0):.2f}%"),
    ]
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        
        # Color code gain/loss
        if "Gain/Loss" in label and isinstance(value, str):
            if data.get('total_gain_loss', 0) >= 0:
                ws[f'B{row}'].font = Font(color="00B050")  # Green
            else:
                ws[f'B{row}'].font = Font(color="FF0000")  # Red
        
        row += 1
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20


def _create_transactions_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create transactions sheet with detailed transaction history."""
    ws = wb.create_sheet("Transactions")
    
    # Headers
    headers = [
        "Date", "Symbol", "Company", "Units", "Price/Unit",
        "Total Invested", "Current Price", "Current Value", 
        "Gain/Loss", "Gain/Loss %", "Transaction ID"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    positions = data.get('positions', [])
    for row, position in enumerate(positions, 2):
        # Convert Decimal to float for Excel
        units = float(position.get('units', 0))
        price_per_unit = float(position.get('price_per_unit', 0))
        total_investment = float(position.get('total_investment', 0))
        current_price = float(position.get('current_price', 0))
        current_value = float(position.get('current_value', 0))
        gain_loss = float(position.get('unrealized_gain_loss', 0))
        gain_loss_pct = float(position.get('unrealized_gain_loss_pct', 0))
        
        ws.cell(row, 1, position.get('transaction_date', ''))
        ws.cell(row, 2, position.get('symbol', ''))
        ws.cell(row, 3, position.get('company_name', ''))
        ws.cell(row, 4, units)
        ws.cell(row, 5, price_per_unit)
        ws.cell(row, 6, total_investment)
        ws.cell(row, 7, current_price)
        ws.cell(row, 8, current_value)
        ws.cell(row, 9, gain_loss)
        ws.cell(row, 10, gain_loss_pct)
        ws.cell(row, 11, position.get('transaction_id', ''))
        
        # Format currency columns
        for col in [5, 6, 7, 8, 9]:
            ws.cell(row, col).number_format = '$#,##0.00'
        
        # Format percentage
        ws.cell(row, 10).number_format = '0.00%'
        ws.cell(row, 10).value = gain_loss_pct / 100
        
        # Color code gain/loss
        if gain_loss >= 0:
            ws.cell(row, 9).font = Font(color="00B050")
            ws.cell(row, 10).font = Font(color="00B050")
        else:
            ws.cell(row, 9).font = Font(color="FF0000")
            ws.cell(row, 10).font = Font(color="FF0000")
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_positions_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create positions sheet with current active positions only."""
    ws = wb.create_sheet("Active Positions")
    
    # Headers
    headers = [
        "Symbol", "Company", "Units", "Avg Cost", "Current Price",
        "Market Value", "Gain/Loss", "Return %"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # All positions from DynamoDB are active (status='ACTIVE')
    positions = data.get('positions', [])
    
    # Data rows
    for row, position in enumerate(positions, 2):
        # Convert Decimal to float for Excel
        units = float(position.get('units', 0))
        price_per_unit = float(position.get('price_per_unit', 0))
        current_price = float(position.get('current_price', 0))
        current_value = float(position.get('current_value', 0))
        gain_loss = float(position.get('unrealized_gain_loss', 0))
        gain_loss_pct = float(position.get('unrealized_gain_loss_pct', 0))
        
        ws.cell(row, 1, position.get('symbol', ''))
        ws.cell(row, 2, position.get('company_name', ''))
        ws.cell(row, 3, units)
        ws.cell(row, 4, price_per_unit)
        ws.cell(row, 5, current_price)
        ws.cell(row, 6, current_value)
        ws.cell(row, 7, gain_loss)
        ws.cell(row, 8, gain_loss_pct)
        
        # Format currency
        for col in [4, 5, 6, 7]:
            ws.cell(row, col).number_format = '$#,##0.00'
        
        # Format percentage
        ws.cell(row, 8).number_format = '0.00%'
        ws.cell(row, 8).value = gain_loss_pct / 100
        
        # Color code
        if gain_loss >= 0:
            ws.cell(row, 7).font = Font(color="00B050")
            ws.cell(row, 8).font = Font(color="00B050")
        else:
            ws.cell(row, 7).font = Font(color="FF0000")
            ws.cell(row, 8).font = Font(color="FF0000")
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_forecasts_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create forecasts sheet comparing predictions vs. actual performance."""
    ws = wb.create_sheet("Forecasts")
    
    # Headers
    headers = [
        "Symbol", "Purchase Date", "Purchase Price", 
        "Forecast Target", "Forecast Days", "Expected Return %",
        "Current Price", "Actual Return %", "Forecast Accuracy"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Filter positions with forecasts
    forecasted_positions = [
        p for p in data.get('positions', []) 
        if p.get('forecast_target_price') is not None
    ]
    
    # Data rows
    for row, position in enumerate(forecasted_positions, 2):
        # Convert Decimal to float for calculations
        purchase_price = float(position.get('price_per_unit', 0))
        forecast_target = float(position.get('forecast_target_price', 0))
        current_price = float(position.get('current_price', 0))
        actual_return = float(position.get('unrealized_gain_loss_pct', 0))
        
        # Calculate expected return
        expected_return = ((forecast_target - purchase_price) / purchase_price * 100) if purchase_price > 0 else 0
        
        # Calculate forecast accuracy
        if expected_return != 0:
            accuracy = (1 - abs(actual_return - expected_return) / abs(expected_return)) * 100
            accuracy = max(0, min(100, accuracy))  # Clamp between 0-100
        else:
            accuracy = 0
        
        ws.cell(row, 1, position.get('symbol', ''))
        ws.cell(row, 2, position.get('transaction_date', ''))
        ws.cell(row, 3, purchase_price)
        ws.cell(row, 4, forecast_target)
        ws.cell(row, 5, int(position.get('forecast_timeframe_days', 0)) if position.get('forecast_timeframe_days') else '')
        ws.cell(row, 6, expected_return)
        ws.cell(row, 7, current_price)
        ws.cell(row, 8, actual_return)
        ws.cell(row, 9, accuracy)
        
        # Format currency
        for col in [3, 4, 7]:
            ws.cell(row, col).number_format = '$#,##0.00'
        
        # Format percentages
        for col in [6, 8, 9]:
            ws.cell(row, col).number_format = '0.00%'
            ws.cell(row, col).value = ws.cell(row, col).value / 100
        
        # Color code accuracy
        if accuracy >= 80:
            ws.cell(row, 9).font = Font(color="00B050")  # Green
        elif accuracy >= 60:
            ws.cell(row, 9).font = Font(color="FFC000")  # Orange
        else:
            ws.cell(row, 9).font = Font(color="FF0000")  # Red
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
